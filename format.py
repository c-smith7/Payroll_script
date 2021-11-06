# This file will hold the openpyxl code.

import pandas as pd
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles.borders import Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle, Alignment, Border, numbers
from openpyxl.formatting.rule import CellIsRule
from transform import transform_script


def format_script(file):
    # get variables from transform.py
    transform_script(file)
    wellness_payroll = transform_script.wellness
    cei_payroll = transform_script.cei
    file_path = transform_script.file_path

    # Instantiate workbook, and activate worksheet for wellness payroll
    wb = Workbook()
    ws_wellness = wb.active
    ws_wellness.title = 'Wellness Payroll'


    # create secondary worksheet for CEI payroll
    ws_cei = wb.create_sheet('CEI Payroll')


    # read in dataframes from transform.py
    for r in dataframe_to_rows(wellness_payroll, index=False, header=True):
        ws_wellness.append(r)


    for r in dataframe_to_rows(cei_payroll, index=False, header=True):
        ws_cei.append(r)


    # insert row above row 1 for title
    ws_wellness.insert_rows(0, amount=1)
    ws_cei.insert_rows(0, amount=1)


    # edit column widths
    wellness_col_widths = [21, 11.5, 11.5, 16, 19, 9.5]
    cei_col_widths = [21, 11.5, 11.5, 16, 23]

    for i, width in enumerate(wellness_col_widths):
        ws_wellness.column_dimensions[get_column_letter(i+1)].width = width

    for i, width in enumerate(cei_col_widths):
        ws_cei.column_dimensions[get_column_letter(i+1)].width = width


    # format column headers
    col_format = NamedStyle(name='col_format')
    col_format.font = Font(bold=True)
    col_format.fill = PatternFill(fgColor='d8d8d9', fill_type='solid')
    wb.add_named_style(col_format)

    for cell in ws_wellness['2:2']:
        cell.style = 'col_format'

    for cell in ws_cei['2:2']:
        cell.style = 'col_format'


    # Add payroll titles to each worksheet and format
    ws_wellness.merge_cells('A1:F1')
    ws_cei.merge_cells('A1:E1')

    wellness_title = ws_wellness['A1']
    cei_title = ws_cei['A1']

    wellness_title.value = 'Wellness Physician Care Payroll'
    cei_title.value = 'CEI Real Estate Payroll'

    title_format = NamedStyle(name='title_format')
    title_format.font = Font(bold=True, size=13)
    title_format.fill = PatternFill(fgColor='b7b7b7', fill_type='solid')
    title_format.alignment = Alignment(horizontal='center')
    wb.add_named_style(title_format)

    wellness_title.style = 'title_format' 
    cei_title.style = 'title_format'


    # Conditional formatting for overtime hours
    last_row_wellness = len(ws_wellness['A'])
    last_row_cei = len(ws_cei['A'])

    red_fill = PatternFill(bgColor='e7b8ae', fill_type='solid')
    ws_wellness.conditional_formatting.add(f'C3:C{last_row_wellness}', CellIsRule(operator='greaterThan', formula=['80'], fill=red_fill))
    ws_cei.conditional_formatting.add(f'C3:C{last_row_cei}', CellIsRule(operator='greaterThan', formula=['80'], fill=red_fill))


    # Total payroll cell format
    total_cell_wellness = f'C{last_row_wellness+1}'
    total_cell_cei = f'C{last_row_cei+1}'

    ws_wellness[total_cell_wellness].value = 'Total Payroll'
    ws_wellness[total_cell_wellness].font = Font(bold=True)
    ws_cei[total_cell_cei].value = 'Total Payroll'
    ws_cei[total_cell_cei].font = Font(bold=True)

    ws_wellness[f'D{last_row_wellness}'].border = Border(bottom=Side(style='double'))
    ws_cei[f'D{last_row_cei}'].border = Border(bottom=Side(style='double'))


    # Total payroll cell formula
    ws_wellness[f'D{last_row_wellness+1}'] = f'=SUM(D3:D{last_row_wellness})'
    ws_cei[f'D{last_row_cei+1}'] = f'=SUM(D3:D{last_row_cei})'


    # Set number format to currency
    for row in ws_wellness.iter_rows(min_row=3, max_row=20, max_col=2, min_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'

    for row in ws_wellness.iter_rows(min_row=3, max_row=21, max_col=4, min_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'

    for row in ws_cei.iter_rows(min_row=3, max_row=25, max_col=2, min_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'

    for row in ws_cei.iter_rows(min_row=3, max_row=25, max_col=4, min_col=4):
        for cell in row:
            cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'


    # get payroll date to use as file name and save.
    summary_df = pd.read_csv(file_path, header=1)
    payroll_date = summary_df.columns[1].replace('/', '-').replace(' To ', '_')
    save_path = f'C:/Users/mcmco/Desktop/QuikMed/Payroll/Final/Payroll_{payroll_date}.xlsx'
    wb.save(save_path)

#format_script()